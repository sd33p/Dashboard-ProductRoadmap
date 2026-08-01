#!/usr/bin/env python3
"""
Build the DW-SFTIES roadmap dashboard from an Excel workbook.

    python scripts/build_dashboard.py [--config config.yml] [--excel path] [--out path]

Reads config.yml, optionally overrides the release fields from a "Dashboard Info"
sheet inside the workbook, then writes a single self-contained HTML file.
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sys
from pathlib import Path

import openpyxl
import yaml

ROOT = Path(__file__).resolve().parent.parent
TEMPLATE = Path(__file__).resolve().parent / "template.html"

# Keys accepted in the workbook's "Dashboard Info" sheet, mapped to config paths.
INFO_SHEET_KEYS = {
    "current release": ("release", "current_release"),
    "current release uat date": ("release", "current_release_uat_date"),
    "release in development": ("release", "release_in_development"),
    "next release uat date": ("release", "next_release_uat_date"),
    "future release label": ("release", "future_release_label"),
    "dashboard title": ("dashboard", "title"),
    "as of label": ("dashboard", "as_of_label"),
    "release notes url": ("dashboard", "release_notes_url"),
    "intro text": ("dashboard", "intro_text"),
}


# --------------------------------------------------------------------------- #
# helpers
# --------------------------------------------------------------------------- #
def norm(value) -> str:
    """Excel cell -> trimmed string with collapsed whitespace."""
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.strftime("%d %B %Y")
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return re.sub(r"\s+", " ", str(value)).strip()


def is_blank_row(cells: list[str]) -> bool:
    return all(c == "" or c == "-" for c in cells)


def die(msg: str) -> "NoReturn":  # noqa: F821
    print(f"ERROR: {msg}", file=sys.stderr)
    raise SystemExit(1)


# --------------------------------------------------------------------------- #
# config
# --------------------------------------------------------------------------- #
def load_config(path: Path) -> dict:
    if not path.exists():
        die(f"config file not found: {path}")
    cfg = yaml.safe_load(path.read_text(encoding="utf-8-sig")) or {}
    for section in ("source", "release", "dashboard", "statuses", "table", "theme", "output"):
        if section not in cfg:
            die(f"config.yml is missing the '{section}' section")
    if not cfg["statuses"]:
        die("config.yml defines no statuses")
    return cfg


# Environment overrides, applied last. Used by the "Run workflow" button so the
# release wording can be changed without editing a file.
ENV_KEYS = {
    "DASH_CURRENT_RELEASE": ("release", "current_release"),
    "DASH_UAT_DATE": ("release", "current_release_uat_date"),
    "DASH_DEV_RELEASE": ("release", "release_in_development"),
    "DASH_NEXT_UAT_DATE": ("release", "next_release_uat_date"),
    "DASH_AS_OF": ("dashboard", "as_of_label"),
}


def apply_env(cfg: dict) -> list[str]:
    import os

    applied = []
    for env_name, (section, key) in ENV_KEYS.items():
        value = norm(os.environ.get(env_name, ""))
        if value:
            cfg[section][key] = value
            applied.append(f"{env_name} -> {section}.{key} = {value}")
    return applied


def apply_info_sheet(wb, cfg: dict) -> list[str]:
    """
    Let a two-column "Dashboard Info" sheet in the workbook override config.yml,
    so the release wording can be updated in the same file you upload.
    Column A = label, column B = value.
    """
    name = cfg["source"].get("info_sheet_name") or ""
    if not name or name not in wb.sheetnames:
        return []
    ws, applied = wb[name], []
    for row in ws.iter_rows(min_col=1, max_col=2, values_only=True):
        label, value = norm(row[0]).lower().rstrip(":"), norm(row[1] if len(row) > 1 else "")
        target = INFO_SHEET_KEYS.get(label)
        if target and value:
            cfg[target[0]][target[1]] = value
            applied.append(f"{label} = {value}")
    return applied


# --------------------------------------------------------------------------- #
# status classification
# --------------------------------------------------------------------------- #
class StatusMap:
    """Maps a cell's raw text to one of the configured statuses."""

    def __init__(self, statuses: list[dict]):
        self.statuses = statuses
        self.lookup: dict[str, dict] = {}
        for st in statuses:
            for token in [st["symbol"]] + list(st.get("match") or []):
                self.lookup[str(token).strip().lower()] = st

    def classify(self, raw: str) -> dict | None:
        key = raw.strip().lower()
        if not key or key in ("n/a", "na", "-"):
            return None
        return self.lookup.get(key)


def overall_status(row_statuses: list[dict | None], statuses: list[dict]) -> str:
    """
    Roll a row's layer statuses into one value:
      any layer under development -> in_development
      else any layer planned      -> planned
      else at least one complete  -> complete
    """
    keys = {s["key"] for s in row_statuses if s}
    order = [st["key"] for st in statuses]
    for key in ("in_development", "planned"):
        if key in keys:
            return key
    if "complete" in keys:
        return "complete"
    for key in order:
        if key in keys:
            return key
    return "none"


# --------------------------------------------------------------------------- #
# read the workbook
# --------------------------------------------------------------------------- #
def read_roadmap(cfg: dict, excel_path: Path):
    if not excel_path.exists():
        die(f"Excel file not found: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=False)
    applied = apply_info_sheet(wb, cfg) + apply_env(cfg)

    sheet_name = cfg["source"].get("sheet_name") or wb.sheetnames[0]
    if sheet_name not in wb.sheetnames:
        die(f"sheet '{sheet_name}' not in workbook. Available: {', '.join(wb.sheetnames)}")
    ws = wb[sheet_name]

    header_row = int(cfg["source"].get("header_row", 1))
    headers = [norm(c.value) for c in ws[header_row]]
    while headers and headers[-1] == "":
        headers.pop()
    if not headers:
        die(f"no column headers found on row {header_row} of '{sheet_name}'")

    tbl = cfg["table"]
    group_cols = [c for c in tbl["group_columns"] if c in headers]
    status_cols = [c for c in tbl["status_columns"] if c in headers]
    latest_col = tbl["latest_release_column"]
    past_col = tbl["past_releases_column"]

    missing = [c for c in tbl["status_columns"] if c not in headers]
    if missing:
        print(f"  note: status columns not found in the sheet, skipped: {', '.join(missing)}")

    columns = []
    for name in headers:
        if name in group_cols:
            ctype, level = "group", group_cols.index(name)
        elif name in status_cols:
            ctype, level = "status", None
        elif name == latest_col:
            ctype, level = "latest", None
        elif name == past_col:
            ctype, level = "past", None
        else:
            ctype, level = "text", None
        columns.append({"name": name, "type": ctype, "groupLevel": level, "choices": []})

    smap = StatusMap(cfg["statuses"])
    current = norm(cfg["release"]["current_release"])
    future_label = norm(cfg["release"]["future_release_label"]).lower()
    theme = cfg["theme"]
    fill_by_key = {st["key"]: st for st in cfg["statuses"]}

    # Rule 4: the current release is orange (reusing the "under development"
    # colour). Rule 5: the future-release label is yellow (the "planned" colour).
    dev_st, planned_st = fill_by_key.get("in_development", {}), fill_by_key.get("planned", {})
    release_fills = {
        "current": {
            "label": f"Current release {current} (in UAT)",
            "fill": dev_st.get("fill", theme["btn_orange"]),
            "text": dev_st.get("text", "#8a4b00"),
        },
        "future": {
            "label": "Planned for a future release",
            "fill": planned_st.get("fill", "#ffffc3"),
            "text": planned_st.get("text", "#7a6a00"),
        },
    }

    def release_key(value: str):
        v = value.strip().lower()
        if not v:
            return None
        if current and v == current.lower():
            return "current"
        if future_label and v == future_label:
            return "future"
        return None

    rows, skipped = [], 0
    idx = {name: i for i, name in enumerate(headers)}

    for r in range(header_row + 1, ws.max_row + 1):
        cells = [norm(ws.cell(r, c).value) for c in range(1, len(headers) + 1)]
        if is_blank_row(cells):
            skipped += 1
            continue

        status_by_index, row_statuses = {}, []
        for name in status_cols:
            i = idx[name]
            st = smap.classify(cells[i])
            if st:
                status_by_index[str(i)] = st["key"]
                row_statuses.append(st)

        rel_value = cells[idx[latest_col]] if latest_col in idx else ""
        row = {
            "cells": cells,
            "status": status_by_index,
            "overall": overall_status(row_statuses, cfg["statuses"]),
        }
        rkey = release_key(rel_value)
        if rkey:
            row["rel"] = rkey
        rows.append(row)

    # Filter dropdown choices: use a <select> for low-cardinality columns.
    for i, col in enumerate(columns):
        values = sorted({r["cells"][i] for r in rows if r["cells"][i]},
                        key=lambda s: (len(s) > 12, s))
        if col["type"] in ("status", "latest") or len(values) <= 60:
            col["choices"] = values

    stats = {
        "total": len(rows),
        "by_status": {st["key"]: sum(1 for r in rows if r["overall"] == st["key"])
                      for st in cfg["statuses"]},
    }
    stats["pct_complete"] = round(
        stats["by_status"].get("complete", 0) / max(stats["total"], 1) * 100, 1
    )

    return columns, rows, stats, skipped, applied, release_fills


# --------------------------------------------------------------------------- #
# render
# --------------------------------------------------------------------------- #
def legend_html(cfg: dict) -> str:
    parts = []
    for st in cfg["statuses"]:
        parts.append(
            '<div class="legend-item">'
            f'<span class="swatch" style="background:{st["fill"]};color:{st.get("text", "#000")}">'
            f'{st["symbol"]}</span>'
            f'<span> = {st["label"]}</span>'
            f'<p class="desc">{st.get("description", "")}</p>'
            "</div>"
        )
    return "\n      ".join(parts)


def render(cfg: dict, columns, rows, stats, excel_path: Path, release_fills: dict) -> str:
    tpl = TEMPLATE.read_text(encoding="utf-8")
    dash, rel, theme, tbl = cfg["dashboard"], cfg["release"], cfg["theme"], cfg["table"]

    as_of = norm(dash.get("as_of_label", "auto"))
    if as_of.lower() in ("auto", ""):
        as_of = dt.date.today().strftime("%B %Y")

    payload = {
        "columns": columns,
        "rows": rows,
        "stats": stats,
        "statuses": [
            {
                "key": s["key"],
                "label": s["label"],
                "symbol": s["symbol"],
                "fill": s["fill"],
                "text": s.get("text", "#000"),
            }
            for s in cfg["statuses"]
        ],
        "releaseFills": release_fills,
        "settings": {
            "simplified_hides_empty_columns": bool(tbl.get("simplified_hides_empty_columns", True)),
            "simplified_hides_past_releases": bool(tbl.get("simplified_hides_past_releases", True)),
            "export_basename": re.sub(r"[^A-Za-z0-9]+", "-",
                                      f"{dash['title']} {as_of}").strip("-").lower(),
        },
    }

    tokens = {
        "__PAGE_TITLE__": f"{dash['title']} — {as_of}",
        "__HEADING__": f"{dash['title']} - {as_of}",
        "__ORG__": dash.get("organization", ""),
        "__INTRO__": norm(dash.get("intro_text", "")),
        "__LEGEND__": legend_html(cfg),
        "__CURRENT_RELEASE__": norm(rel["current_release"]),
        "__UAT_DATE__": norm(rel["current_release_uat_date"]),
        "__DEV_RELEASE__": norm(rel["release_in_development"]),
        "__NEXT_UAT_DATE__": norm(rel["next_release_uat_date"]),
        "__RELEASE_NOTES_URL__": dash.get("release_notes_url", "#"),
        "__SOURCE_FILE__": excel_path.name,
        "__ROW_COUNT__": str(len(rows)),
        "__BUILD_STAMP__": dt.datetime.now(dt.timezone.utc).strftime("%d %b %Y %H:%M UTC"),
        "__C_BLUE__": theme["blue"],
        "__C_PAGE_BLUE__": theme["page_blue"],
        "__C_INK__": theme["ink"],
        "__C_GROUP_GREY__": theme["group_grey"],
        "__C_NA_TEXT__": theme["na_text"],
        "__C_BTN_ORANGE__": theme["btn_orange"],
        "__C_BTN_PURPLE__": theme["btn_purple"],
        "__C_BTN_GREEN__": theme["btn_green"],
    }

    # Payload first so a stray token inside the data is never substituted.
    html = tpl.replace("__PAYLOAD_JSON__",
                       json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
                       .replace("</", "<\\/"))
    for token, value in tokens.items():
        html = html.replace(token, str(value))

    leftover = re.findall(r"__[A-Z_]+__", html)
    if leftover:
        die(f"template tokens were left unfilled: {sorted(set(leftover))}")
    return html


# --------------------------------------------------------------------------- #
def main() -> None:
    ap = argparse.ArgumentParser(description="Build the roadmap dashboard.")
    ap.add_argument("--config", default=str(ROOT / "config.yml"))
    ap.add_argument("--excel", default=None, help="override source.excel_file")
    ap.add_argument("--out", default=None, help="override output.html_file")
    args = ap.parse_args()

    cfg = load_config(Path(args.config))
    excel_path = Path(args.excel or (ROOT / cfg["source"]["excel_file"]))
    out_path = Path(args.out or (ROOT / cfg["output"]["html_file"]))

    print(f"Reading {excel_path}")
    columns, rows, stats, skipped, applied, release_fills = read_roadmap(cfg, excel_path)

    if applied:
        print("  release info taken from the workbook's info sheet:")
        for line in applied:
            print(f"    - {line}")

    if not rows:
        die("no data rows found — check source.sheet_name and source.header_row")

    html = render(cfg, columns, rows, stats, excel_path, release_fills)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(html, encoding="utf-8")

    print(f"  {len(columns)} columns, {len(rows)} rows ({skipped} blank rows skipped)")
    print(f"  current release {cfg['release']['current_release']} highlighted orange, "
          f"'{cfg['release']['future_release_label']}' highlighted yellow")
    for key, count in stats["by_status"].items():
        print(f"  {key}: {count}")
    print(f"Wrote {out_path} ({out_path.stat().st_size / 1024:.0f} KB)")


if __name__ == "__main__":
    main()
