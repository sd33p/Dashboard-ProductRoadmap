#!/usr/bin/env python3
"""
Add (or refresh) the "Dashboard Info" sheet in the roadmap workbook.

That sheet lets you update the release wording in the same file you upload, so
you never have to touch config.yml. Run once; afterwards just edit the yellow
cells in Excel each cycle.

    python scripts/add_config_sheet.py [path/to/roadmap.xlsx]
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import yaml
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parent.parent
SHEET = "Dashboard Info"

BLUE = "FF0150B7"
INPUT_FILL = PatternFill("solid", fgColor="FFFFF2CC")   # yellow = you edit these
NOTE_FILL = PatternFill("solid", fgColor="FFF2F2F2")
THIN = Side(style="thin", color="FFBFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# (label, config path, help text). Labels must match INFO_SHEET_KEYS in
# build_dashboard.py.
FIELDS = [
    ("Current release", ("release", "current_release"),
     "Latest shipped release. Cells in the Latest Release column matching this turn ORANGE."),
    ("Current release UAT date", ("release", "current_release_uat_date"),
     "Date the current release was promoted to UAT, e.g. 06 June 2026."),
    ("Release in development", ("release", "release_in_development"),
     "Release the team is building now."),
    ("Next release UAT date", ("release", "next_release_uat_date"),
     "When that release is expected in UAT, e.g. August 2026."),
    ("Future release label", ("release", "future_release_label"),
     "Text in the Latest Release column that should turn YELLOW."),
    ("As of label", ("dashboard", "as_of_label"),
     'Month shown in the page title. Leave as "auto" to use the build date.'),
    ("Release notes URL", ("dashboard", "release_notes_url"),
     "Where the View release notes button points."),
]


def main() -> None:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else None
    cfg = yaml.safe_load((ROOT / "config.yml").read_text(encoding="utf-8-sig"))
    if xlsx is None:
        xlsx = ROOT / cfg["source"]["excel_file"]
    if not xlsx.exists():
        print(f"ERROR: {xlsx} not found", file=sys.stderr)
        raise SystemExit(1)

    wb = openpyxl.load_workbook(xlsx)
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET, 0)

    ws["A1"] = "Dashboard release information"
    ws["A1"].font = Font(name="Arial", size=13, bold=True, color="FFFFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=BLUE)
    ws.merge_cells("A1:C1")
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26

    ws["A2"] = ("Edit the yellow cells in column B, save, and upload this workbook. "
                "The dashboard rebuilds itself from these values.")
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="FF555555")
    ws.merge_cells("A2:C2")

    for col, head in zip("ABC", ("Setting", "Value", "What it controls")):
        c = ws[f"{col}4"]
        c.value = head
        c.font = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=BLUE)
        c.border = BORDER
        c.alignment = Alignment(vertical="center")

    row = 5
    for label, (section, key), note in FIELDS:
        ws.cell(row, 1, label).font = Font(name="Arial", size=10, bold=True)
        value = ws.cell(row, 2, str(cfg[section][key]))
        value.font = Font(name="Arial", size=10)
        value.fill = INPUT_FILL
        value.alignment = Alignment(horizontal="left")
        note_cell = ws.cell(row, 3, note)
        note_cell.font = Font(name="Arial", size=9, color="FF555555")
        note_cell.fill = NOTE_FILL
        note_cell.alignment = Alignment(wrap_text=True, vertical="center")
        for col in (1, 2, 3):
            ws.cell(row, col).border = BORDER
        ws.row_dimensions[row].height = 30
        row += 1

    ws.cell(row + 1, 1, "Yellow cells = the only cells you need to change.").font = Font(
        name="Arial", size=9, italic=True, color="FF7A6A00")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 78
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    wb.save(xlsx)
    print(f"Added '{SHEET}' sheet to {xlsx}")
    print("Edit the yellow cells in column B each release cycle.")


if __name__ == "__main__":
    main()
